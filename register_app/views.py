from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from .forms import RegistrationForm
from .models import Registration
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.utils import timezone  # Import timezone module to handle time zones

# Registration form view - handles both form display and submission
def form(request):
    # Check if the request is a form submission (POST method)
    if request.method == 'POST':
        # Initialize form with submitted data
        form = RegistrationForm(request.POST)
        # Validate the form data
        if form.is_valid():
            # Save the form data to the database if valid
            form.save()
            # Add success message to be displayed to the user
            messages.success(request, 'Registration successful!')
            # Redirect to the same page to prevent form resubmission
            return redirect('form')
    else:
        # For GET requests, create an empty form
        form = RegistrationForm()
    
    # Render the template with the form (empty for GET, with errors for invalid POST)
    return render(request, 'form.html', {'form': form})

# Excel export view - generates an Excel file with all registration data
def export_to_excel(request):
    # Create a new Excel workbook and get the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Registrations"
    
    # Define column headers for the Excel file
    headers = [
        "ID", "First Name", "Last Name", "Email", 
        "Date of Birth", "Phone Number", "Registration Date"
    ]
    
    # Add and style headers in the first row
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Fetch all registration records from the database
    registrations = Registration.objects.all()
    
    # Add each registration record to the worksheet (starting from row 2)
    for row_num, registration in enumerate(registrations, 2):
        # Note: Password is intentionally excluded for security reasons
        worksheet.cell(row=row_num, column=1).value = registration.id
        worksheet.cell(row=row_num, column=2).value = registration.first_name
        worksheet.cell(row=row_num, column=3).value = registration.last_name
        worksheet.cell(row=row_num, column=4).value = registration.email
        worksheet.cell(row=row_num, column=5).value = registration.date_of_birth
        worksheet.cell(row=row_num, column=6).value = registration.phone_number
        
        # Remove timezone information from created_at to make it compatible with Excel
        if registration.created_at:
            created_at_naive = timezone.make_naive(registration.created_at) if timezone.is_aware(registration.created_at) else registration.created_at
            worksheet.cell(row=row_num, column=7).value = created_at_naive
        else:
            worksheet.cell(row=row_num, column=7).value = None
    
    # Auto-adjust column widths for better readability
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate unique filename with timestamp to prevent overwriting
    filename = f"registrations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Set up HTTP response with appropriate headers for file download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Write Excel file content to the HTTP response
    workbook.save(response)
    
    return response

# View to display all registration information
def all_registrations(request):
    # Fetch all registration records from the database
    registrations = Registration.objects.all().order_by('-created_at')
    
    # Render the template with all registrations data
    return render(request, 'registrations_list.html', {'registrations': registrations})

# View to delete a registration record
def delete_registration(request, id):
    # Get the registration object or return 404 if not found
    registration = get_object_or_404(Registration, pk=id)
    
    if request.method == 'POST':
        # Delete the registration record
        registration.delete()
        # Add success message
        messages.success(request, f'Registration for {registration.first_name} {registration.last_name} has been deleted.')
        # Redirect to registrations list
        return redirect('all_registrations')
    
    # For GET requests, show confirmation page
    return render(request, 'confirm_delete.html', {'registration': registration})

# View to delete all registration records
def delete_all_registrations(request):
    if request.method == 'POST':
        # Get the count before deletion
        count = Registration.objects.count()
        # Delete all registration records
        Registration.objects.all().delete()
        # Add success message
        messages.success(request, f'All {count} registrations have been deleted.')
        # Redirect to registrations list
        return redirect('all_registrations')
    
    # For GET requests, show confirmation page
    return render(request, 'confirm_delete_all.html')
